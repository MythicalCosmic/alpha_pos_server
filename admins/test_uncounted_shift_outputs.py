"""Manager-facing shift outputs must never turn an omitted count into a loss."""
from datetime import timedelta
from io import BytesIO

import pytest
from django.utils import timezone
from openpyxl import load_workbook


pytestmark = pytest.mark.django_db


def _ended_shift(cashier, *, counted_methods):
    from base.models import Shift

    end = timezone.now()
    return Shift.objects.create(
        user=cashier,
        branch_id='branch1',
        start_time=end - timedelta(hours=1),
        end_time=end,
        status=Shift.Status.ENDED,
        settlement_manifest={
            'version': 3,
            'cashier_counted_methods': counted_methods,
        },
    )


def _cash_settlement(shift, *, expected='267000.00'):
    from base.models import Order, OrderPayment
    from cashbox.models import ShiftPaymentTotal

    order = Order.objects.create(
        user=shift.user,
        cashier=shift.user,
        branch_id=shift.branch_id,
        display_id=Order.objects.count() + 1,
        status=Order.Status.COMPLETED,
        is_paid=True,
        payment_method=Order.PaymentMethod.CASH,
        subtotal=expected,
        total_amount=expected,
        paid_at=shift.start_time + timedelta(minutes=30),
    )
    OrderPayment.objects.create(
        order=order,
        method=Order.PaymentMethod.CASH,
        amount=expected,
    )
    return ShiftPaymentTotal.objects.create(
        shift=shift,
        branch_id=shift.branch_id,
        method='CASH',
        expected_amount=expected,
        counted_amount='0.00',
        confirmed_amount='0.00',
        difference=f'-{expected}',
    )


def test_handover_json_nulls_unsubmitted_count_and_shortage(cashier_user):
    from admins.services.shift_analytics_service import shift_handover_report

    shift = _ended_shift(cashier_user, counted_methods=[])
    _cash_settlement(shift)

    report = shift_handover_report(shift)

    assert report['settlement'] == [{
        'method': 'CASH',
        'expected': '267000.00',
        'frozen_expected': '267000.00',
        'expected_source': 'FROZEN_MATCHED',
        'counted': None,
        'cashier_count_submitted': False,
        'cashier_count_status': 'UNCOUNTED',
        'confirmed': None,
        'manager_confirmed': False,
        'confirmation_source': None,
        'confirmation_difference': None,
        'difference': None,
        'frozen_difference': '-267000.00',
        'difference_source': 'FROZEN_MATCHED',
        'status': 'UNCOUNTED',
        'reconciled': False,
        'shift_reconciled': False,
    }]


def test_handover_json_preserves_explicit_zero_count(cashier_user):
    from admins.services.shift_analytics_service import shift_handover_report

    shift = _ended_shift(cashier_user, counted_methods=['CASH'])
    _cash_settlement(shift)

    cash = shift_handover_report(shift)['settlement'][0]

    assert cash['status'] == 'COUNTED'
    assert cash['counted'] == '0.00'
    assert cash['cashier_count_submitted'] is True
    assert cash['cashier_count_status'] == 'COUNTED'
    assert cash['confirmed'] is None
    assert cash['confirmation_difference'] is None
    assert cash['difference'] == '-267000.00'


def test_manager_confirmation_preserves_missing_cashier_count_in_json_and_xlsx(
    cashier_user,
    admin_user,
):
    from admins.services.shift_analytics_service import shift_handover_report
    from admins.services.workbook_export_service import (
        build_shift_report_workbook,
    )
    from core.shifts.service import ShiftService

    shift = _ended_shift(cashier_user, counted_methods=[])
    # Model a legacy close without the explicit blind-count marker. The
    # manager may reconcile it, but that must not invent a cashier count.
    shift.settlement_manifest = {}
    shift.save(update_fields=['settlement_manifest'])
    _cash_settlement(shift)
    admin_user.branch_id = shift.branch_id
    admin_user.save(update_fields=['branch_id'])

    response, status = ShiftService.reconcile(
        shift.id,
        '267000.00',
        '',
        admin_user.id,
        actor=admin_user,
    )

    assert status in (200, 201), response
    report = shift_handover_report(shift)
    cash = report['settlement'][0]
    assert cash['status'] == 'CONFIRMED'
    assert cash['cashier_count_status'] == 'UNCOUNTED'
    assert cash['counted'] is None
    assert cash['difference'] is None
    assert cash['confirmed'] == '267000.00'
    assert cash['confirmation_difference'] == '0.00'

    sheet = load_workbook(
        BytesIO(build_shift_report_workbook(report)),
        data_only=False,
    )['Settlement']
    assert [sheet.cell(5, column).value for column in range(1, 7)] == [
        'CASH', 267000, None, 267000, None, 'CONFIRMED',
    ]


def test_workbook_blanks_only_uncounted_placeholders():
    from admins.services.workbook_export_service import (
        build_shift_report_workbook,
    )

    report = {
        'shift': {'shift_id': 7},
        'cashier': {'name': 'Test Cashier'},
        'settlement': [
            {
                'method': 'CASH',
                'status': 'UNCOUNTED',
                'expected': '267000.00',
                'counted': '0.00',
                'confirmed': '0.00',
                'difference': '-267000.00',
            },
            {
                'method': 'HUMO',
                'status': 'COUNTED',
                'expected': '100.00',
                'counted': '90.00',
                'confirmed': '0.00',
                'difference': '-10.00',
            },
            {
                'method': 'UZCARD',
                'status': 'CONFIRMED',
                'expected': '100.00',
                'counted': '100.00',
                'confirmed': '100.00',
                'difference': '0.00',
            },
        ],
    }

    content = build_shift_report_workbook(report)
    sheet = load_workbook(
        BytesIO(content), data_only=False,
    )['Settlement']

    assert [sheet.cell(4, column).value for column in range(1, 7)] == [
        'Method', 'Expected', 'Counted', 'Confirmed', 'Difference', 'Status',
    ]
    assert [sheet.cell(5, column).value for column in range(1, 7)] == [
        'CASH', 267000, None, None, None, 'UNCOUNTED',
    ]
    assert [sheet.cell(6, column).value for column in range(1, 7)] == [
        'HUMO', 100, 90, None, -10, 'COUNTED',
    ]
    assert [sheet.cell(7, column).value for column in range(1, 7)] == [
        'UZCARD', 100, 100, 100, 0, 'CONFIRMED',
    ]
