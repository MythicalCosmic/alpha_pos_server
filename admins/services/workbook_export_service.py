"""Native XLSX exports for the admin dashboard and shift handover report.

The analytics services remain the source of truth.  This module only renders
their already-filtered dictionaries, so a spreadsheet cannot drift from the
numbers displayed by the JSON endpoints.
"""
import re
from collections.abc import Mapping
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

_NAVY = '17324D'
_TEAL = '18A7A0'
_WHITE = 'FFFFFF'
_GRID = Side(style='thin', color='D9E2E8')
_FORMULA_PREFIXES = ('=', '+', '-', '@')
_PLAIN_NUMBER = re.compile(r'^-?\d+(?:\.\d+)?$')
_ILLEGAL_XML_CHARS = re.compile(
    r'[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF\uFFFE\uFFFF]'
)


def _excel_value(value, *, numeric=False):
    """Return an Excel scalar while preventing formula injection.

    Numeric conversion is opt-in because identifiers and user-controlled names
    may legitimately look numeric (for example ``"001"``).  Converting every
    digit-only string would lose leading zeroes and could round long text IDs.
    """
    if value is None:
        return ''
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, (Decimal, float)):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()

    text = str(value)
    if numeric:
        stripped = text.strip()
        number = None
        # Never feed arbitrary/exponential user text to Decimal/int.  Besides
        # preserving identifiers, the length and grammar bound prevents values
        # such as ``1e999999999`` from becoming a spreadsheet-export DoS.
        if len(stripped) <= 40 and _PLAIN_NUMBER.fullmatch(stripped):
            try:
                number = Decimal(stripped)
            except (InvalidOperation, TypeError, ValueError):
                number = None
        if number is not None and number.is_finite() and stripped:
            if number == number.to_integral_value():
                return int(number)
            return float(number)

    text = _ILLEGAL_XML_CHARS.sub('', text)
    if text.lstrip().startswith(_FORMULA_PREFIXES):
        text = "'" + text
    # Excel cells are limited to 32,767 characters.  Bound notes/reasons here
    # so a single oversized text value cannot break or bloat the workbook.
    return text[:32767]


def _label(value):
    return str(value).replace('_', ' ').replace('.', ' / ').strip().title()


def _write_title(sheet, title, subtitle='', *, width=8):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = sheet.cell(1, 1, _excel_value(title))
    cell.font = Font(size=18, bold=True, color=_WHITE)
    cell.fill = PatternFill('solid', fgColor=_NAVY)
    cell.alignment = Alignment(vertical='center')
    sheet.row_dimensions[1].height = 30
    if subtitle:
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        sub = sheet.cell(2, 1, _excel_value(subtitle))
        sub.font = Font(size=10, italic=True, color='526575')
    sheet.sheet_view.showGridLines = False


def _style_header(cell):
    cell.font = Font(bold=True, color=_WHITE)
    cell.fill = PatternFill('solid', fgColor=_TEAL)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(bottom=_GRID)


def _write_table(
        sheet, start_row, headers, rows, *, freeze=True, numeric_columns=()):
    headers = list(headers)
    rows = list(rows)
    numeric_columns = set(numeric_columns)
    for col, header in enumerate(headers, start=1):
        _style_header(sheet.cell(start_row, col, _label(header)))

    for row_offset, row in enumerate(rows, start=1):
        if isinstance(row, Mapping):
            values = [row.get(header) for header in headers]
        else:
            values = list(row)
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(
                start_row + row_offset,
                col,
                _excel_value(value, numeric=headers[col - 1] in numeric_columns),
            )
            cell.border = Border(bottom=_GRID)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = '#,##0.##'
        if row_offset % 2 == 0:
            for col in range(1, len(headers) + 1):
                sheet.cell(start_row + row_offset, col).fill = PatternFill(
                    'solid', fgColor='F6FAFC',
                )

    end_row = start_row + max(len(rows), 1)
    if headers:
        sheet.auto_filter.ref = (
            f'A{start_row}:{get_column_letter(len(headers))}{end_row}'
        )
    if freeze:
        sheet.freeze_panes = f'A{start_row + 1}'
    _fit_columns(sheet, len(headers), end_row)
    return end_row


def _fit_columns(sheet, column_count, end_row):
    for column in range(1, column_count + 1):
        values = (
            sheet.cell(row, column).value
            for row in range(1, end_row + 1)
        )
        length = max((len(str(value or '')) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column)].width = min(
            max(length + 2, 12), 42,
        )


def _flatten(mapping, prefix=''):
    rows = []
    for key, value in (mapping or {}).items():
        path = f'{prefix}.{key}' if prefix else str(key)
        if isinstance(value, Mapping):
            rows.extend(_flatten(value, path))
        elif isinstance(value, (list, tuple)):
            rows.append((path, ', '.join(str(item) for item in value)))
        else:
            rows.append((path, value))
    return rows


def _sheet_with_table(
        workbook, name, title, headers, rows, *, subtitle='', width=None,
        numeric_columns=()):
    sheet = workbook.create_sheet(name)
    width = width or max(len(headers), 2)
    _write_title(sheet, title, subtitle, width=width)
    _write_table(
        sheet, 4, headers, rows, numeric_columns=numeric_columns,
    )
    return sheet


def _dynamic_headers(rows, preferred=()):
    headers = list(preferred)
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return headers


def _workbook_bytes(workbook):
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _generated_label(generated_at=None):
    generated_at = generated_at or timezone.now()
    if timezone.is_aware(generated_at):
        generated_at = timezone.localtime(generated_at)
    return generated_at.isoformat(timespec='seconds')


def _metric_value(metric, value):
    """Keep human text as text; make analytics scalar strings numeric."""
    lowered = metric.lower()
    text_suffixes = (
        '.name', '.user_name', '.status', '_time', '_at', '.notes',
        '.reconciled_by',
    )
    if lowered.endswith(text_suffixes):
        return _excel_value(value)
    return _excel_value(value, numeric=True)


def build_dashboard_workbook(data, *, filters=None, generated_at=None):
    """Render the exact ``dashboard_service.get_range`` result to XLSX."""
    filters = filters or {}
    date_range = data.get('range') or {}
    date_from = date_range.get('from') or ''
    date_to = date_range.get('to') or ''
    subtitle = f'Business dates {date_from} to {date_to}'

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = 'Alpha POS'
    workbook.properties.title = 'Dashboard export'
    workbook.properties.subject = _excel_value(subtitle)

    summary = workbook.create_sheet('Summary')
    _write_title(summary, 'Alpha POS Dashboard', subtitle, width=4)
    summary_rows = [
        ('Generated at', _generated_label(generated_at)),
        ('Business date from', date_from),
        ('Business date to', date_to),
        ('Time from', filters.get('tod_from') or 'All day'),
        ('Time to', filters.get('tod_to') or 'All day'),
        ('Net revenue (UZS)', _excel_value(data.get('revenue', 0), numeric=True)),
        ('Gross revenue (UZS)', _excel_value(data.get('gross_revenue', 0), numeric=True)),
        ('Refund amount (UZS)', _excel_value(data.get('refund_amount', 0), numeric=True)),
        ('Orders', data.get('orders', 0)),
        ('Paid orders', data.get('paid_orders', 0)),
        ('Refunded orders', data.get('refunded_orders', 0)),
        ('Cancelled orders', data.get('cancelled', 0)),
        ('Units sold', data.get('units_sold', 0)),
    ]
    _write_table(summary, 4, ('metric', 'value'), summary_rows, freeze=False)

    payment = data.get('payment_breakdown') or {}
    payment_rows = [
        {'method': method, 'amount_uzs': amount}
        for method, amount in payment.items()
        if method != 'card_detail'
    ]
    _sheet_with_table(
        workbook,
        'Payments',
        'Payment Breakdown',
        ('method', 'amount_uzs'),
        payment_rows,
        subtitle=subtitle,
        numeric_columns=('amount_uzs',),
    )
    card_detail_rows = [
        {'acquirer': method, 'amount_uzs': amount}
        for method, amount in (payment.get('card_detail') or {}).items()
    ]
    _sheet_with_table(
        workbook,
        'Card Details',
        'Card Acquirer Detail',
        ('acquirer', 'amount_uzs'),
        card_detail_rows,
        subtitle=f'{subtitle} / components of the card total',
        numeric_columns=('amount_uzs',),
    )

    product_headers = (
        'product_id', 'product_name', 'quantity', 'revenue',
        'gross_quantity', 'refunded_quantity', 'gross_revenue',
        'refund_amount',
    )
    _sheet_with_table(
        workbook,
        'Top Products',
        'Top Products',
        product_headers,
        data.get('top_products') or [],
        subtitle=subtitle,
        numeric_columns={
            'product_id', 'quantity', 'revenue', 'gross_quantity',
            'refunded_quantity', 'gross_revenue', 'refund_amount',
        },
    )

    category_headers = (
        'category_id', 'category', 'quantity', 'revenue',
        'gross_quantity', 'refunded_quantity', 'gross_revenue',
        'refund_amount',
    )
    _sheet_with_table(
        workbook,
        'Categories',
        'Category Performance',
        category_headers,
        data.get('category_stats') or [],
        subtitle=subtitle,
        numeric_columns={
            'category_id', 'quantity', 'revenue', 'gross_quantity',
            'refunded_quantity', 'gross_revenue', 'refund_amount',
        },
    )
    return _workbook_bytes(workbook)


def build_shift_report_workbook(report, *, generated_at=None):
    """Render one canonical ``shift_handover_report`` result to XLSX."""
    shift = report.get('shift') or {}
    shift_id = shift.get('shift_id') or ''
    cashier = report.get('cashier') or {}
    subtitle = f'Shift {shift_id} / {cashier.get("name") or "Unknown cashier"}'

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = 'Alpha POS'
    workbook.properties.title = f'Shift {shift_id} handover report'
    workbook.properties.subject = _excel_value(subtitle)

    summary = workbook.create_sheet('Summary')
    _write_title(summary, 'Alpha POS Shift Report', subtitle, width=4)
    summary_rows = [
        ('generated_at', _generated_label(generated_at)),
        ('cashier.id', cashier.get('id')),
        ('cashier.name', cashier.get('name')),
        ('receipt_count', report.get('receipt_count', 0)),
        ('peak_hour', report.get('peak_hour')),
        *_flatten(shift, 'shift'),
        *_flatten(report.get('best_seller') or {}, 'best_seller'),
    ]
    summary_rows = [
        (metric, _metric_value(metric, value))
        for metric, value in summary_rows
    ]
    _write_table(summary, 4, ('metric', 'value'), summary_rows, freeze=False)

    sections = (
        (
            'Settlement',
            'Tender Settlement',
            report.get('settlement') or [],
            ('method', 'expected', 'counted', 'confirmed', 'difference'),
            {'expected', 'counted', 'confirmed', 'difference'},
        ),
        (
            'Cash Expenses',
            'Cash Expenses',
            report.get('cash_expenses') or [],
            ('category', 'total', 'count'),
            {'total', 'count'},
        ),
        (
            'Receipts',
            'Receipts',
            report.get('receipts') or [],
            (
                'order_id', 'display_id', 'status', 'order_type', 'is_paid',
                'payment_method', 'total_amount', 'discount_amount',
                'discount_percent', 'line_items', 'units', 'created_at',
                'paid_at',
            ),
            {
                'order_id', 'display_id', 'total_amount', 'discount_amount',
                'discount_percent', 'line_items', 'units',
            },
        ),
        (
            'Refunds',
            'Refunds',
            report.get('refunds') or [],
            (
                'refund_id', 'order_id', 'amount', 'cash', 'card', 'payme',
                'refunded_at', 'reason',
            ),
            {'refund_id', 'order_id', 'amount', 'cash', 'card', 'payme'},
        ),
        (
            'Products',
            'Product Performance',
            report.get('products') or [],
            (
                'product_id', 'name', 'units_sold', 'times_sold',
                'times_refunded', 'revenue',
            ),
            {
                'product_id', 'units_sold', 'times_sold', 'times_refunded',
                'revenue',
            },
        ),
    )
    for name, title, rows, preferred, numeric_columns in sections:
        headers = _dynamic_headers(rows, preferred)
        _sheet_with_table(
            workbook,
            name,
            title,
            headers,
            rows,
            subtitle=subtitle,
            numeric_columns=numeric_columns,
        )

    distribution = report.get('distribution') or {}
    for key, name, title, preferred in (
        ('by_hour', 'Hourly', 'Hourly Distribution', ('hour', 'orders', 'revenue')),
        ('by_date', 'Daily', 'Daily Distribution', ('date', 'orders', 'revenue')),
    ):
        rows = distribution.get(key) or []
        headers = _dynamic_headers(rows, preferred)
        _sheet_with_table(
            workbook,
            name,
            title,
            headers,
            rows,
            subtitle=subtitle,
            numeric_columns={
                header for header in headers
                if header not in {'date', 'weekday', 'label'}
            },
        )

    return _workbook_bytes(workbook)
