"""Browser-download contracts for dashboard and shift-report XLSX exports."""
import secrets
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.test import Client
from django.utils import timezone
from openpyxl import load_workbook

from base.repositories.session import SessionRepository


pytestmark = pytest.mark.django_db
BUSINESS_DATE = date(2026, 7, 15)


def _auth_token(user):
    from base.models import Session

    payload = secrets.token_hex(32)
    Session.objects.create(
        user_id=user,
        ip_address='127.0.0.1',
        payload=SessionRepository.hash_token(payload),
        expires_at=timezone.now() + timedelta(hours=1),
    )
    return payload


def _auth(user):
    return {'HTTP_AUTHORIZATION': f'Bearer {_auth_token(user)}'}


def _product(name='Export Burger', category_name='Export Category'):
    from base.models import Category, Product

    suffix = Category.objects.count() + 1
    category = Category.objects.create(
        name=category_name,
        slug=f'export-category-{suffix}',
    )
    return Product.objects.create(
        name=name,
        category=category,
        price=Decimal('100'),
    )


def _paid_order(
        regular_user, cashier, product, when, *, amount='100', method='CASH'):
    from base.models import Order, OrderItem, OrderPayment

    order = Order.objects.create(
        user=regular_user,
        cashier=cashier,
        branch_id='branch1',
        status=Order.Status.COMPLETED,
        order_type=Order.OrderType.HALL,
        is_paid=True,
        payment_method=method,
        paid_at=when,
        subtotal=Decimal(amount),
        total_amount=Decimal(amount),
        display_id=Order.objects.count() + 1,
        order_number=Order.objects.count() + 1,
    )
    Order.objects.filter(pk=order.pk).update(created_at=when)
    OrderItem.objects.create(
        order=order,
        product=product,
        quantity=1,
        price=Decimal(amount),
        original_price=Decimal(amount),
    )
    OrderPayment.objects.create(
        order=order,
        method=method,
        amount=Decimal(amount),
    )
    order.refresh_from_db()
    return order


def _rows_by_first_column(sheet):
    return {
        row[0]: row[1]
        for row in sheet.iter_rows(min_row=5, values_only=True)
        if row[0] is not None
    }


def test_dashboard_export_is_filter_consistent_xlsx_attachment(
        admin_user, cashier_user, regular_user):
    from base.services.business_day import range_window

    start, end = range_window(BUSINESS_DATE, BUSINESS_DATE)
    product = _product(
        name='\x0b=HYPERLINK("https://invalid", "Burger")',
        category_name='001',
    )
    _paid_order(
        regular_user, cashier_user, product,
        start + timedelta(hours=4), amount='125000',
    )
    _paid_order(
        regular_user, cashier_user, product,
        start + timedelta(hours=5), amount='50000', method='UZCARD',
    )
    _paid_order(
        regular_user, cashier_user, product,
        start + timedelta(hours=6), amount='25000', method='HUMO',
    )
    _paid_order(
        regular_user, cashier_user, product,
        end + timedelta(hours=1), amount='900000',
    )

    client = Client()
    auth = _auth(admin_user)
    response = client.get(
        '/api/admins/dashboard/export',
        {
            'from': BUSINESS_DATE.isoformat(),
            'to': BUSINESS_DATE.isoformat(),
            'tod_from': 'invalid',
        },
        **auth,
    )

    assert response.status_code == 200
    assert response['Content-Type'] == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    assert response.content.startswith(b'PK')
    assert 'alpha-pos-dashboard-2026-07-15-to-2026-07-15.xlsx' in (
        response['Content-Disposition']
    )
    assert response['Cache-Control'] == 'private, no-store'
    assert response['X-Export-Count'] == '3'
    assert 'Content-Disposition' in response['Access-Control-Expose-Headers']

    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == [
        'Summary', 'Payments', 'Card Details', 'Top Products', 'Categories',
    ]
    summary = _rows_by_first_column(workbook['Summary'])
    assert summary['Time from'] == 'All day'
    assert summary['Net revenue (UZS)'] == 200000
    assert summary['Gross revenue (UZS)'] == 200000
    assert summary['Orders'] == 3
    assert summary['Paid orders'] == 3

    payments = _rows_by_first_column(workbook['Payments'])
    assert payments['cash'] == 125000
    assert payments['card'] == 75000
    assert sum(payments.values()) == summary['Net revenue (UZS)']
    card_details = _rows_by_first_column(workbook['Card Details'])
    assert sum(card_details.values()) == payments['card']
    product_name = workbook['Top Products']['B5']
    assert product_name.data_type == 's'
    assert product_name.value.startswith("'=")
    assert workbook['Categories']['B5'].value == '001'

    filtered_query = {
        'from': BUSINESS_DATE.isoformat(),
        'to': BUSINESS_DATE.isoformat(),
        'tod_from': '08:00',
        'tod_to': '08:30',
    }
    filtered_json = client.get(
        '/api/admins/dashboard', filtered_query, **auth,
    ).json()['data']
    filtered_export = client.get(
        '/api/admins/dashboard/export', filtered_query, **auth,
    )
    filtered_workbook = load_workbook(BytesIO(filtered_export.content))
    filtered_summary = _rows_by_first_column(filtered_workbook['Summary'])
    filtered_payments = _rows_by_first_column(filtered_workbook['Payments'])
    assert filtered_summary['Net revenue (UZS)'] == int(
        filtered_json['revenue']
    )
    # Legacy date+clock inputs now resolve to one exact continuous interval.
    # No seeded receipt falls in 08:00-08:30.
    assert filtered_summary['Orders'] == filtered_json['orders'] == 0
    assert filtered_payments['card'] == int(
        filtered_json['payment_breakdown']['card']
    ) == 0


def test_dashboard_export_auth_method_and_empty_range(
        admin_user, cashier_user):
    client = Client()
    path = (
        '/api/admins/dashboard/export?from=2020-01-01&to=2020-01-01'
        '&tod_from=22:00&tod_to=02:00'
    )
    assert client.get(path).status_code == 401
    assert client.get(path, **_auth(cashier_user)).status_code == 403
    assert client.post(path, **_auth(admin_user)).status_code == 405

    response = client.get(path, **_auth(admin_user))
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    summary = _rows_by_first_column(workbook['Summary'])
    assert summary['Orders'] == 0
    assert summary['Time from'] == '22:00'
    assert summary['Time to'] == '02:00'
    assert workbook['Top Products'].max_row == 4


def test_shift_report_export_matches_json_and_preserves_ownership(
        admin_user, cashier_user, other_cashier_user, regular_user,
        monkeypatch):
    from base.models import Shift, User
    from base.security.hashing import hash_password
    from base.services.business_day import range_window

    start, _ = range_window(BUSINESS_DATE, BUSINESS_DATE)
    shift = Shift.objects.create(
        user=cashier_user,
        branch_id='branch1',
        start_time=start + timedelta(hours=1),
        end_time=start + timedelta(hours=5),
        status=Shift.Status.ENDED,
        total_orders=1,
        total_revenue=Decimal('150000'),
        cash_collected=Decimal('150000'),
    )
    product = _product(name='@Shift Formula', category_name='Shift Category')
    order = _paid_order(
        regular_user, cashier_user, product,
        shift.start_time + timedelta(minutes=30), amount='150000',
    )
    path = f'/api/admins/analytics/shifts/{shift.id}/report/export'
    json_path = f'/api/admins/analytics/shifts/{shift.id}/report'
    client = Client()

    auth = _auth(admin_user)
    json_response = client.get(json_path, **auth)
    response = client.get(path, **auth)
    assert json_response.status_code == response.status_code == 200
    report = json_response.json()['data']

    assert response.content.startswith(b'PK')
    assert f'alpha-pos-shift-{shift.id}-report-2026-07-15.xlsx' in (
        response['Content-Disposition']
    )
    assert response['X-Export-Count'] == '1'
    assert response['Cache-Control'] == 'private, no-store'

    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == [
        'Summary', 'Settlement', 'Cash Expenses', 'Receipts', 'Refunds',
        'Products', 'Hourly', 'Daily',
    ]
    summary = _rows_by_first_column(workbook['Summary'])
    assert summary['shift.money.revenue'] == int(
        Decimal(report['shift']['money']['revenue'])
    )
    assert summary['shift.money.cash'] == 150000
    assert summary['best_seller.name'] == "'@Shift Formula"
    assert workbook['Receipts']['A5'].value == order.id
    assert workbook['Receipts']['G5'].value == 150000
    product_name = workbook['Products']['B5']
    assert product_name.data_type == 's'
    assert product_name.value == "'@Shift Formula"

    assert client.get(path, **_auth(cashier_user)).status_code == 200
    assert client.get(path, **_auth(other_cashier_user)).status_code == 403
    manager = User.objects.create(
        first_name='Export',
        last_name='Manager',
        email='export.manager@test.local',
        password=hash_password('managerpass'),
        role=User.RoleChoices.MANAGER,
        status=User.UserStatus.ACTIVE,
    )
    assert client.get(path, **_auth(manager)).status_code == 200
    assert client.get(path).status_code == 401
    assert client.get(
        '/api/admins/analytics/shifts/999999/report/export',
        **_auth(admin_user),
    ).status_code == 404
    assert client.post(path, **_auth(admin_user)).status_code == 405
    monkeypatch.setattr(
        'admins.views.analytics_views._shift_export_receipt_count',
        lambda _shift: 5001,
    )
    too_large = client.get(path, **_auth(admin_user))
    assert too_large.status_code == 413
    assert too_large.json()['max_receipts'] == 5000
    Shift.objects.filter(pk=shift.pk).update(is_deleted=True)
    assert client.get(path, **_auth(admin_user)).status_code == 404
